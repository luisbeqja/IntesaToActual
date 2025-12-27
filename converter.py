"""
CSV/XLSX Converter for Intesa SanPaolo bank exports to Actual Budget format.
"""

import csv
import io
from typing import TextIO, Union, BinaryIO

# Input column names (Italian)
INPUT_COLUMNS = {
    'date': 'Data',
    'payee': 'Operazione',
    'notes': 'Dettagli',
    'amount': 'Importo',
}

# Output column names (English - Actual Budget format)
OUTPUT_COLUMNS = ['Account', 'Date', 'Payee', 'Notes', 'Category', 'Amount', 'Split_Amount', 'Cleared']

# Fixed account name for output
ACCOUNT_NAME = 'Intesa SanPaolo'


def _find_header_row_csv(lines: list[str]) -> int:
    """Find the header row index in CSV content."""
    for i, line in enumerate(lines):
        if 'Data' in line and 'Operazione' in line and 'Dettagli' in line:
            return i
    raise ValueError("Could not find header row in input file")


def _find_header_row_xlsx(sheet) -> int:
    """Find the header row index in XLSX sheet."""
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_str = ' '.join(str(cell) if cell else '' for cell in row)
        if 'Data' in row_str and 'Operazione' in row_str and 'Dettagli' in row_str:
            return row_idx
    raise ValueError("Could not find header row in input file")


def _get_column_indices(header_row: list) -> dict:
    """Get column indices from header row."""
    indices = {}
    for i, cell in enumerate(header_row):
        cell_str = str(cell).strip() if cell else ''
        if cell_str == INPUT_COLUMNS['date']:
            indices['date'] = i
        elif cell_str == INPUT_COLUMNS['payee']:
            indices['payee'] = i
        elif cell_str == INPUT_COLUMNS['notes']:
            indices['notes'] = i
        elif cell_str == INPUT_COLUMNS['amount']:
            indices['amount'] = i
    return indices


def transform_xlsx(input_file: Union[str, BinaryIO], output_file: Union[str, TextIO, None] = None) -> str:
    """
    Transform an Intesa SanPaolo XLSX export to Actual Budget format.
    
    Args:
        input_file: Path to input XLSX file or file-like object
        output_file: Path to output CSV file or file-like object (optional)
    
    Returns:
        The transformed CSV content as a string
    """
    from openpyxl import load_workbook
    
    # Load workbook (don't use read_only mode as it may not read all rows correctly)
    if isinstance(input_file, str):
        wb = load_workbook(input_file, data_only=True)
    else:
        wb = load_workbook(input_file, data_only=True)
    
    sheet = wb.active
    
    # Find header row
    header_row_idx = _find_header_row_xlsx(sheet)
    
    # Get all rows as list for easier processing
    all_rows = list(sheet.iter_rows(values_only=True))
    header_row = all_rows[header_row_idx - 1]  # Convert to 0-indexed
    
    # Get column indices
    col_indices = _get_column_indices(header_row)
    
    # Transform rows
    output_buffer = io.StringIO()
    writer = csv.DictWriter(output_buffer, fieldnames=OUTPUT_COLUMNS)
    writer.writeheader()
    
    for row in all_rows[header_row_idx:]:  # Start after header
        # Skip empty rows
        if not any(row):
            continue
        
        # Get values with safe indexing
        date_val = row[col_indices.get('date', 0)] if col_indices.get('date') is not None else ''
        payee_val = row[col_indices.get('payee', 0)] if col_indices.get('payee') is not None else ''
        notes_val = row[col_indices.get('notes', 0)] if col_indices.get('notes') is not None else ''
        amount_val = row[col_indices.get('amount', 0)] if col_indices.get('amount') is not None else ''
        
        # Format date if it's a datetime object
        if hasattr(date_val, 'strftime'):
            date_val = date_val.strftime('%d/%m/%Y')
        
        output_row = {
            'Account': ACCOUNT_NAME,
            'Date': str(date_val) if date_val else '',
            'Payee': str(payee_val) if payee_val else '',
            'Notes': str(notes_val) if notes_val else '',
            'Category': '',
            'Amount': str(amount_val) if amount_val else '',
            'Split_Amount': '',
            'Cleared': '',
        }
        writer.writerow(output_row)
    
    wb.close()
    result = output_buffer.getvalue()
    
    # Write to output file if specified
    if output_file is not None:
        if isinstance(output_file, str):
            with open(output_file, 'w', encoding='utf-8', newline='') as f:
                f.write(result)
        else:
            output_file.write(result)
    
    return result


def transform_csv(input_file: Union[str, TextIO], output_file: Union[str, TextIO, None] = None) -> str:
    """
    Transform an Intesa SanPaolo CSV export to Actual Budget format.
    
    Args:
        input_file: Path to input CSV file or file-like object
        output_file: Path to output CSV file or file-like object (optional)
    
    Returns:
        The transformed CSV content as a string
    """
    # Handle input
    if isinstance(input_file, str):
        with open(input_file, 'r', encoding='utf-8') as f:
            content = f.read()
    else:
        content = input_file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8')
    
    # Split into lines
    lines = content.splitlines()
    
    # Find the header row
    header_row_index = _find_header_row_csv(lines)
    
    # Parse CSV starting from header row
    csv_content = '\n'.join(lines[header_row_index:])
    reader = csv.DictReader(io.StringIO(csv_content))
    
    # Transform rows
    output_buffer = io.StringIO()
    writer = csv.DictWriter(output_buffer, fieldnames=OUTPUT_COLUMNS)
    writer.writeheader()
    
    for row in reader:
        output_row = {
            'Account': ACCOUNT_NAME,
            'Date': row.get(INPUT_COLUMNS['date'], ''),
            'Payee': row.get(INPUT_COLUMNS['payee'], ''),
            'Notes': row.get(INPUT_COLUMNS['notes'], ''),
            'Category': '',
            'Amount': row.get(INPUT_COLUMNS['amount'], ''),
            'Split_Amount': '',
            'Cleared': '',
        }
        writer.writerow(output_row)
    
    result = output_buffer.getvalue()
    
    # Write to output file if specified
    if output_file is not None:
        if isinstance(output_file, str):
            with open(output_file, 'w', encoding='utf-8', newline='') as f:
                f.write(result)
        else:
            output_file.write(result)
    
    return result


def transform_file(input_file: Union[str, TextIO, BinaryIO], output_file: Union[str, TextIO, None] = None, filename: str = None) -> str:
    """
    Transform an Intesa SanPaolo export (CSV or XLSX) to Actual Budget format.
    Automatically detects file type from extension or filename parameter.
    
    Args:
        input_file: Path to input file or file-like object
        output_file: Path to output CSV file or file-like object (optional)
        filename: Original filename (used for type detection when input_file is a file object)
    
    Returns:
        The transformed CSV content as a string
    """
    # Determine file type
    if isinstance(input_file, str):
        file_ext = input_file.lower().split('.')[-1]
    elif filename:
        file_ext = filename.lower().split('.')[-1]
    else:
        raise ValueError("Cannot determine file type. Provide filename parameter for file objects.")
    
    if file_ext == 'xlsx':
        return transform_xlsx(input_file, output_file)
    elif file_ext == 'csv':
        return transform_csv(input_file, output_file)
    else:
        raise ValueError(f"Unsupported file type: {file_ext}. Use .csv or .xlsx files.")


def main():
    """CLI entry point for the converter."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python converter.py <input.csv|input.xlsx> [output.csv]")
        print("\nTransforms Intesa SanPaolo CSV/XLSX exports to Actual Budget format.")
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        result = transform_file(input_path, output_path)
        
        if output_path:
            print(f"Converted successfully! Output saved to: {output_path}")
        else:
            print(result)
    except FileNotFoundError:
        print(f"Error: File not found: {input_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
